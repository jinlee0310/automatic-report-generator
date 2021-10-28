import openpyxl
from query_generator import query_generator, make_txt, get_filename
from openpyxl.styles import Font, PatternFill, Alignment
from time import sleep


def create_sheet_contents(sheet, file_name, title, sheet_name, column_name, query):
    # setting styles
    ft = Font(name="휴먼명조", size=9, bold=True)
    alignment = Alignment(horizontal="center", vertical="center")

    # A column
    sheet['A1'] = "진단테이블"
    sheet['A2'] = "진단파일명"
    sheet['A3'] = "진단컬럼"
    sheet['A4'] = "진단항목명"
    sheet['A5'] = "업무규칙 내용"
    sheet['A6'] = "진단 SQL"+"\n"+"(오류 추출 SQL)"

    for a in range(len(sheet['A'])):
        sheet.cell(row=a+1, column=1).font = ft
        sheet.cell(row=a+1, column=1).alignment = alignment

    # B column
    sheet['B1'] = file_name
    sheet['B2'] = title
    sheet['B3'] = sheet_name
    sheet['B4'] = column_name
    sheet['B6'] = query


def create_sheet(filename, queries):
    for c in range(2, len(col)+1):
        if sheet1.cell(row=c, column=6).value == 'Y':
            try:
                sheet_name = sheet1.cell(row=c, column=4).value
                sheet = wb.create_sheet(sheet_name)
                title = sheet1.cell(row=c, column=3).value
                column_name = sheet1.cell(row=c, column=5).value
                query = queries[sheet_name]
                create_sheet_contents(sheet, filename, title,
                                      sheet_name, column_name, query)
            except:
                pass


if __name__ == "__main__":
    print("보고서 생성기를 실행합니다...")
    sleep(0.5)
    filename = get_filename()
    xlsx_name = '{}_SQL진단결과보고서.xlsx'.format(filename)
    # 이름 틀릴 경우
    try:
        wb = openpyxl.load_workbook(xlsx_name)
        sheet1 = wb.active
        sheet1.title = "01.컬럼목록"
        # print(sheet1)
        col = sheet1['F']
        queries = query_generator()
        print("생성된 쿼리를 저장하시겠습니까? y/n")
        answer = input()
        if answer == "y" or answer == "Y":
            make_txt(filename)
        create_sheet(filename, queries)

        # 보고서 저장
        try:
            wb.save(xlsx_name)
            sleep(0.5)
            print("==========================================")
            print("보고서가 저장되었습니다.")
        # permission deny
        except PermissionError:
            print("보고서 파일을 닫고 프로그램을 재실행해주세요.")
        # save실패할 경우
        except:
            print("보고서 저장에 실패하였습니다.")

    except FileNotFoundError:
        print('엑셀 파일이 존재하지 않습니다.')
